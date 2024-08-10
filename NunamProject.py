#!/usr/bin/env python
# coding: utf-8

# In[3]:


#1st Excel File 5329
import openpyxl

# Define variable to load the dataframe
excel1 = "C:\\Users\\GVJai\\Desktop\\Project\\Nunam\\5329.xlsx"
dataframe = openpyxl.load_workbook(excel1)

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)


# In[120]:


dataframe["Cycle_67_3_1"]


# In[223]:
import pandas as pd

sheet=dataframe["Cycle_67_3_1"]
data = sheet.values
columns = next(data)[0:]  # Get the header row
df = pd.DataFrame(data, columns=columns)

print(df) 


# In[224]:


import mysql.connector
def soh():
    mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

    cursor = mydb.cursor()
    drop_query = """DROP TABLE IF EXISTS soh"""
    cursor.execute(drop_query)
    mydb.commit()


    create_query = """CREATE TABLE IF NOT EXISTS soh(
        Channel INT, 
        Total_Cycle INT PRIMARY KEY,
        Capacity_of_charge FLOAT,
        Capacity_of_discharge FLOAT,
        Life_Cycle FLOAT
    )"""
    cursor.execute(create_query)
    mydb.commit()

    for index, row in df.iterrows():
            insert_query = '''INSERT INTO soh(Channel, Total_Cycle, Capacity_of_charge, Capacity_of_discharge, Life_Cycle) 
                               VALUES(%s, %s, %s, %s, %s)'''
            values = (
                int(row["Channel"]),
                int(row["ToTal of Cycle"]),
                float(row["Capacity of charge(mAh)"]),
                float(row["Capacity of discharge(mAh)"]),
                float(row["Cycle Life(%)"])
            )

            cursor.execute(insert_query, values)
            mydb.commit()


# In[225]:


soh()


# In[170]:


sheet1=dataframe["Detail_67_3_1"]
data1 = sheet1.values
columns1 = next(data1)[0:]  # Get the header row
df1 = pd.DataFrame(data1, columns=columns1)

print(df1.columns) 


# In[228]:


import mysql.connector

def data1_detail():
    # Connect to MySQL database
    mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

    cursor = mydb.cursor()

    # Drop the table if it exists
    drop_query = """DROP TABLE IF EXISTS data1_detail"""
    cursor.execute(drop_query)
    mydb.commit()

    # Create the table
    create_query = """CREATE TABLE IF NOT EXISTS data1_detail(
        Record_Index INT PRIMARY KEY,
        Status VARCHAR(100),
        JumpTo FLOAT,
        Cycle FLOAT,
        Step FLOAT,
        Cur FLOAT,
        Voltage FLOAT,
        Capacity FLOAT,
        Energy Float,
        Relative_Time VARCHAR(255),
        Absolute_Time VARCHAR(255)
    )"""
    cursor.execute(create_query)
    mydb.commit()

    # Insert data into the table
    insert_query = '''INSERT INTO data1_detail(
        Record_Index, Status, JumpTo, Cycle, Step, Cur, Voltage, Capacity, Energy, Relative_Time, Absolute_Time
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'''

    for index, row in df1.iterrows():
        values = (
            row["Record Index"], row["Status"], row["JumpTo"], row["Cycle"], row["Step"],
            row["Cur(mA)"], row["Voltage(V)"], row["CapaCity(mAh)"], row["Energy(mWh)"], row["Relative Time(h:min:s.ms)"], row["Absolute Time"]
        )
        try:
            cursor.execute(insert_query, values)
            mydb.commit()
        except mysql.connector.Error as e:
            print(f"Error inserting data: {e}")
            mydb.rollback()

    # Close the cursor and connection
    cursor.close()
    mydb.close()


# In[229]:


data1_detail()


# In[167]:


sheet2=dataframe["DetailTemp_67_3_1"]
data2 = sheet2.values
columns2 = next(data2)[0:]  # Get the header row
df1 = pd.DataFrame(data2, columns=columns2)

print(df2.columns) 


# In[168]:


import mysql.connector

def data1_temp():
    # Connect to MySQL database
    mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

    cursor = mydb.cursor()

    # Drop the table if it exists
    drop_query = """DROP TABLE IF EXISTS data1_temp"""
    cursor.execute(drop_query)
    mydb.commit()

    # Create the table
    create_query = """CREATE TABLE IF NOT EXISTS data1_temp(
        Record_Index INT PRIMARY KEY,
        Step_name VARCHAR(100),
        Relative_Time VARCHAR(100),
        Realtime VARCHAR(100),
        Auxiliary_channel FLOAT,
        Gap_of_Temperature FLOAT
    )"""
    cursor.execute(create_query)
    mydb.commit()

    # Insert data into the table
    insert_query = '''INSERT INTO data1_temp(
        Record_Index, Step_name, Relative_Time, Realtime, Auxiliary_channel, Gap_of_Temperature
    ) VALUES (%s, %s, %s, %s, %s, %s)'''

    for index, row in df2.iterrows():
        values = (
            row["Record ID"], row["Step Name"], row["Relative Time(h:min:s.ms)"], row["Realtime"], row["Auxiliary channel TU1 T(°C)"],
            row["Gap of Temperature"]
        )
        try:
            cursor.execute(insert_query, values)
            mydb.commit()
        except mysql.connector.Error as e:
            print(f"Error inserting data: {e}")
            mydb.rollback()

    # Close the cursor and connection
    cursor.close()
    mydb.close()


# In[ ]:


data1_temp()


# In[175]:


get_ipython().system('pip install mysql-connector-python pandas')


# In[226]:


import mysql.connector
import pandas as pd
import matplotlib.pyplot as plt

# Connect to MySQL
mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

# Create a cursor object
cursor = mydb.cursor()

# Define your query
query = "SELECT Total_Cycle, Capacity_of_discharge FROM soh"

# Execute the query
cursor.execute(query)

# Fetch all rows from the executed query
rows = cursor.fetchall()

# Get column names
columns = [desc[0] for desc in cursor.description]

# Create DataFrame
df = pd.DataFrame(rows, columns=columns)

# Close cursor and connection
cursor.close()
mydb.close()

# Perform the necessary calculations
df['Capacity_of_discharge'] = (df['Capacity_of_discharge'] / 3000) * 100

# Create a pie chart using Matplotlib
labels = df['Total_Cycle']
sizes = df['Capacity_of_discharge']
explode = (0.1, 0, 0)  # explode the 1st slice (if needed)

fig1, ax1 = plt.subplots()
ax1.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
        shadow=True, startangle=90)
ax1.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.

plt.title("Capacity of Discharge per Total Cycle")
plt.show()


# In[ ]:





# In[216]:


import mysql.connector
import pandas as pd

# Connect to MySQL
mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

# Create a cursor object
cursor = mydb.cursor()

# Define your query
query = "SELECT Absolute_Time, Voltage FROM data1_detail limit 20"  # Limit to 5 rows for inspection

# Execute the query
cursor.execute(query)

# Fetch all rows from the executed query
rows = cursor.fetchall()

# Get column names
columns = [desc[0] for desc in cursor.description]

# Create DataFrame
df = pd.DataFrame(rows, columns=columns)

# Close cursor and connection
cursor.close()
mydb.close()


# Sort the DataFrame by 'Absolute_Time' to ensure proper plotting
df = df.sort_values('Absolute_Time')

# Create a line graph using Matplotlib
plt.figure(figsize=(10, 6))
plt.plot(df['Absolute_Time'], df['Voltage'], marker='o')

plt.title("Voltage vs Time")
plt.xlabel("Time")
plt.ylabel("Voltage")
plt.grid(True)

plt.xticks(rotation=45)  # Rotate x-axis labels for better readability
plt.tight_layout()  # Adjust layout for better fit

plt.show()


# In[218]:


import mysql.connector
import pandas as pd

# Connect to MySQL
mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

# Create a cursor object
cursor = mydb.cursor()

# Define your query
query = "select Cur, Absolute_Time from data1_detail limit 20"  # Limit to 5 rows for inspection

# Execute the query
cursor.execute(query)

# Fetch all rows from the executed query
rows = cursor.fetchall()

# Get column names
columns = [desc[0] for desc in cursor.description]

# Create DataFrame
df = pd.DataFrame(rows, columns=columns)

# Close cursor and connection
cursor.close()
mydb.close()


# Sort the DataFrame by 'Absolute_Time' to ensure proper plotting
df = df.sort_values('Absolute_Time')

# Create a line graph using Matplotlib
plt.figure(figsize=(10, 6))
plt.plot(df['Absolute_Time'], df['Cur'], marker='o')

plt.title("Cur vs Time")
plt.xlabel("Time")
plt.ylabel("Cur")
plt.grid(True)

plt.xticks(rotation=45)  # Rotate x-axis labels for better readability
plt.tight_layout()  # Adjust layout for better fit

plt.show()


# In[219]:


import mysql.connector
import pandas as pd

# Connect to MySQL
mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

# Create a cursor object
cursor = mydb.cursor()

# Define your query
query = "SELECT Absolute_Time, Capacity FROM data1_detail limit 20"  # Limit to 5 rows for inspection

# Execute the query
cursor.execute(query)

# Fetch all rows from the executed query
rows = cursor.fetchall()

# Get column names
columns = [desc[0] for desc in cursor.description]

# Create DataFrame
df = pd.DataFrame(rows, columns=columns)

# Close cursor and connection
cursor.close()
mydb.close()


# Sort the DataFrame by 'Absolute_Time' to ensure proper plotting
df = df.sort_values('Absolute_Time')

# Create a line graph using Matplotlib
plt.figure(figsize=(10, 6))
plt.plot(df['Absolute_Time'], df['Capacity'], marker='o')

plt.title("Capacity vs Time")
plt.xlabel("Time")
plt.ylabel("Capacity")
plt.grid(True)

plt.xticks(rotation=45)  # Rotate x-axis labels for better readability
plt.tight_layout()  # Adjust layout for better fit

plt.show()


# In[220]:


import mysql.connector
import pandas as pd

# Connect to MySQL
mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

# Create a cursor object
cursor = mydb.cursor()

# Define your query
query = "select Absolute_Time, Auxiliary_channel from data1_detail Join data1_temp on data1_detail.Record_Index=data1_temp.Record_Index limit 20"  # Limit to 5 rows for inspection

# Execute the query
cursor.execute(query)

# Fetch all rows from the executed query
rows = cursor.fetchall()

# Get column names
columns = [desc[0] for desc in cursor.description]

# Create DataFrame
df = pd.DataFrame(rows, columns=columns)

# Close cursor and connection
cursor.close()
mydb.close()


# Sort the DataFrame by 'Absolute_Time' to ensure proper plotting
df = df.sort_values('Absolute_Time')

# Create a line graph using Matplotlib
plt.figure(figsize=(10, 6))
plt.plot(df['Absolute_Time'], df['Auxiliary_channel'], marker='o')

plt.title("Auxiliary_channel vs Time")
plt.xlabel("Time")
plt.ylabel("Auxiliary_channel")
plt.grid(True)

plt.xticks(rotation=45)  # Rotate x-axis labels for better readability
plt.tight_layout()  # Adjust layout for better fit

plt.show()


# In[4]:


# 2nd Excel File
import openpyxl

# Define variable to load the dataframe
excel1 = "C:\\Users\\GVJai\\Desktop\\Project\\Nunam\\5308.xlsx"
dataframe = openpyxl.load_workbook(excel1)

# Define variable to read sheet
dataframe1 = dataframe.active



# In[6]:


import pandas as pd
sheet3=dataframe["Cycle_67_3_5"]
data3 = sheet3.values
columns3 = next(data3)[0:]  # Get the header row
df3 = pd.DataFrame(data3, columns=columns3)

print(df3) 


# In[238]:


import mysql.connector
def soh1():
    mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

    cursor = mydb.cursor()
    drop_query = """DROP TABLE IF EXISTS soh1"""
    cursor.execute(drop_query)
    mydb.commit()


    create_query = """CREATE TABLE IF NOT EXISTS soh1(
        Channel INT, 
        Total_Cycle INT PRIMARY KEY,
        Capacity_of_charge FLOAT,
        Capacity_of_discharge FLOAT,
        Life_Cycle FLOAT
    )"""
    cursor.execute(create_query)
    mydb.commit()

    for index, row in df3.iterrows():
            insert_query = '''INSERT INTO soh1(Channel, Total_Cycle, Capacity_of_charge, Capacity_of_discharge, Life_Cycle) 
                               VALUES(%s, %s, %s, %s, %s)'''
            values = (row["Channel"], row["ToTal of Cycle"], row["Capacity of charge(mAh)"], row["Capacity of discharge(mAh)"], 
                      row["Cycle Life(%)"])

            cursor.execute(insert_query, values)
            mydb.commit()


# In[239]:


soh1()


# In[9]:


sheet4=dataframe["Detail_67_3_5"]
data4 = sheet4.values
columns4 = next(data4)[0:]  # Get the header row
df4 = pd.DataFrame(data4, columns=columns4)

print(df4) 


# In[10]:


df4.columns


# In[11]:


import mysql.connector

def data2_detail():
    # Connect to MySQL database
    mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

    cursor = mydb.cursor()

    # Drop the table if it exists
    drop_query = """DROP TABLE IF EXISTS data2_detail"""
    cursor.execute(drop_query)
    mydb.commit()

    # Create the table
    create_query = """CREATE TABLE IF NOT EXISTS data2_detail(
        Record_Index INT PRIMARY KEY,
        Status VARCHAR(100),
        JumpTo FLOAT,
        Cycle FLOAT,
        Step FLOAT,
        Cur FLOAT,
        Voltage FLOAT,
        Capacity FLOAT,
        Energy Float,
        Relative_Time VARCHAR(255),
        Absolute_Time VARCHAR(255)
    )"""
    cursor.execute(create_query)
    mydb.commit()

    # Insert data into the table
    insert_query = '''INSERT INTO data2_detail(
        Record_Index, Status, JumpTo, Cycle, Step, Cur, Voltage, Capacity, Energy, Relative_Time, Absolute_Time
    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'''

    for index, row in df4.iterrows():
        values = (
            row["Record Index"], row["Status"], row["JumpTo"], row["Cycle"], row["Step"],
            row["Cur(mA)"], row["Voltage(V)"], row["CapaCity(mAh)"], row["Energy(mWh)"], row["Relative Time(h:min:s.ms)"], row["Absolute Time"]
        )
        try:
            cursor.execute(insert_query, values)
            mydb.commit()
        except mysql.connector.Error as e:
            print(f"Error inserting data: {e}")
            mydb.rollback()

    # Close the cursor and connection
    cursor.close()
    mydb.close()


# In[13]:


data2_detail()


# In[14]:


sheet5=dataframe["DetailTemp_67_3_5"]
data5 = sheet5.values
columns5 = next(data5)[0:]  # Get the header row
df5 = pd.DataFrame(data5, columns=columns5)

print(df5.columns) 


# In[15]:


import mysql.connector

def data2_temp():
    # Connect to MySQL database
    mydb = mysql.connector.connect(host="localhost",
                user="jai",
                password="jai123",
                database= "nunam",
                auth_plugin='mysql_native_password',
                charset='utf8mb4'
                )

    cursor = mydb.cursor()

    # Drop the table if it exists
    drop_query = """DROP TABLE IF EXISTS data2_temp"""
    cursor.execute(drop_query)
    mydb.commit()

    # Create the table
    create_query = """CREATE TABLE IF NOT EXISTS data2_temp(
        Record_Index INT PRIMARY KEY,
        Step_name VARCHAR(100),
        Relative_Time VARCHAR(100),
        Realtime VARCHAR(100),
        Auxiliary_channel FLOAT,
        Gap_of_Temperature FLOAT
    )"""
    cursor.execute(create_query)
    mydb.commit()

    # Insert data into the table
    insert_query = '''INSERT INTO data2_temp(
        Record_Index, Step_name, Relative_Time, Realtime, Auxiliary_channel, Gap_of_Temperature
    ) VALUES (%s, %s, %s, %s, %s, %s)'''

    for index, row in df5.iterrows():
        values = (
            row["Record ID"], row["Step Name"], row["Relative Time(h:min:s.ms)"], row["Realtime"], row["Auxiliary channel TU1 T(°C)"],
            row["Gap of Temperature"]
        )
        try:
            cursor.execute(insert_query, values)
            mydb.commit()
        except mysql.connector.Error as e:
            print(f"Error inserting data: {e}")
            mydb.rollback()

    # Close the cursor and connection
    cursor.close()
    mydb.close()


# In[16]:


data2_temp()


# In[ ]:




